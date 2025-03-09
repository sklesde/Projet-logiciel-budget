from os import listdir, path, remove, makedirs
from pandas import read_csv, read_excel, to_datetime, to_numeric, ExcelWriter
import pandas as pd
from datetime import date
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from spire.xls import Workbook
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import shutil
import time
import os
from spire.xls import Workbook
from openpyxl import load_workbook


#%%
   
def ouverture_csv(date):
    try:
        # Recherche du premier fichier CSV dans le répertoire courant
        csv_files = [f for f in listdir() if f.endswith(".csv")]
        
        if len(csv_files) == 0:
            raise FileNotFoundError("Aucun fichier CSV trouvé dans le dossier.")
            """
            Par la suite, si aucun dossier n'est trouvé on renvoiera qu'aucun dossier n'a été trouve
            mais que la catégorisation de Budget excel à été mise à jours.
            """
        elif len(csv_files) > 1:
            raise ValueError("Plusieurs fichiers CSV sont présents dans le dossier.")
        
        path_data= csv_files[0]
        data = read_csv(path_data, encoding="utf-8", sep=";")
        data.to_excel(f".\Stockage CSV Banque\{date}.xlsx")
        print(f"Le CSV rentré à été stocké sous le nom de {date} dans le répertoire Stockage CSV Banque.")
        
        L=(data, path_data)
        return L
    
    except FileNotFoundError as e:
        print(str(e))
        return (0,0)
    except ValueError as e:
        print(str(e))
        return (0,0)
    except Exception as e:
        print(f"Une erreur inattendue s'est produite : {str(e)}")
        return (0,0)

def verifier_dossiers():
    dossier_precedentes_versions = "Précédentes versions"
    dossier_stockage_csv_banque = "Stockage CSV Banque"
    
    # Vérification de l'existence des dossiers
    existe_precedentes_versions = path.exists(dossier_precedentes_versions)
    existe_stockage_csv_banque = path.exists(dossier_stockage_csv_banque)
    
    # Si le dossier "Précédentes versions" n'existe pas, on l'ajoute au compteur et on le crée
    if not existe_precedentes_versions:
        makedirs(dossier_precedentes_versions)
        print('Le dossier "Précédentes versions" à été créé.')
    
    # Si le dossier "Stockage CSV Banque" n'existe pas, on l'ajoute au compteur et on le crée
    if not existe_stockage_csv_banque:
        makedirs(dossier_stockage_csv_banque)
        print('Le dossier "Stockage CSV Banque" à été créé.')
        
    return

def convertisseur_en_chiffre(df, columns):
    """
    Convertit les colonnes spécifiées d'un DataFrame en type float.
    Les valeurs non convertibles sont remplacées par NaN.
    
    :param df: DataFrame pandas
    :param columns: Liste des noms de colonnes à convertir en float
    :return: DataFrame avec les colonnes converties
    """
    for column in columns:
        if column in df.columns:
            # Nettoyer les valeurs en remplaçant les virgules par des points
            # et en supprimant les autres caractères non numériques (comme les symboles)
            df[column] = df[column].str.replace(',', '.', regex=False)  # Remplacer la virgule par le point
            
            # Supprimer les caractères non numériques, à l'exception du point et du signe moins
            df[column] = df[column].str.replace(r'[^\d.-]', '', regex=True)
            
            # Conversion des colonnes en float, avec gestion des erreurs
            df[column] = to_numeric(df[column], errors='coerce')
    
    return df



def generer_id_unique(df, date_col="Date operation", libelle_col="Libelle operation"):
    """
    Génère un ID unique pour chaque transaction en fonction du jour et du libellé d'opération.
    Si plusieurs transactions sont identiques sur un même jour, elles sont numérotées (A1, A2, B1, etc.).
    
    :param df: DataFrame contenant les transactions
    :param date_col: Nom de la colonne contenant la date
    :param libelle_col: Nom de la colonne contenant le libellé d'opération
    :return: DataFrame avec une colonne 'ID' unique
    """
    df = df.copy()
    
    # Assurer que la colonne date est bien au format datetime
    df[date_col] = to_datetime(df[date_col], format='%d/%m/%Y', errors='coerce')
    
    # Convertir la date en format string pour l'utiliser dans l'ID (sans l'heure)
    df["date_str"] = df[date_col].dt.strftime("%Y%m%d")  # Seulement la date sans l'heure


    # Grouper par date et libellé d'opération et générer des indices uniques
    df["ID_count"] = df.groupby(["date_str", libelle_col]).cumcount() + 1

    # Création de l'ID unique
    df['ID'] = df.apply(lambda x: f"{x['date_str']}_{x[libelle_col][:10].upper()}_{x['ID_count']}", axis=1)

    # Vérification des doublons et correction
    unique_ids = set()
    for i in df.index:
        id_val = df.at[i, "ID"]
        while id_val in unique_ids:
            df.at[i, "ID_count"] += 1
            id_val = f"{df.at[i, 'date_str']}_{df.at[i, libelle_col][:10].upper()}_{df.at[i, 'ID_count']}"
        unique_ids.add(id_val)
        df.at[i, "ID"] = id_val

    # Suppression des colonnes temporaires
    df = df.drop(columns=["date_str", "ID_count"])

    return df

def creation_data_cp(data, budget_mensuel_donnees):
    try :
        #Je crée une copie de data pour manipuler les données
        colonnes_a_garder = ["Date operation", "Libelle simplifie", "Libelle operation", "Categorie", "Sous categorie", "Debit", "Credit"]

        data_cp = data[colonnes_a_garder].copy() 
        colonnes=["Debit", "Credit"]
        data_cp = convertisseur_en_chiffre(data_cp, colonnes)

        #Je crée la colonne ID et Classification avec des None à l'intérieur
        data_cp['ID']=None
        data_cp['Classification']=None

        #Je crée les ID uniques
        data_cp = generer_id_unique(data_cp)
        
        
        
        return data_cp
        print('test')
    except:
        print("Pas de fichier CSV détecté, mise à jours des catégories.")
    return 0


def verifier_et_organiser_colonnes(feuille_excel: pd.DataFrame, colonnes_attendues: list) -> pd.DataFrame:
    # Si le DataFrame est vide, créer un DataFrame avec les colonnes spécifiées
    if feuille_excel.empty:
        feuille_excel = pd.DataFrame(columns=colonnes_attendues)
    
    # Réorganiser les colonnes pour correspondre à colonnes_attendues, peu importe l'ordre
    feuille_excel = feuille_excel.reindex(columns=colonnes_attendues)
    
    return feuille_excel



def verification_et_fusion(df1, df2, id_col="ID"):
    # Vérification si df2 est un entier
    if isinstance(df2, int):
        return df1
    
    # Vérification que les ID du df2 ne sont pas déjà dans df1
    ids_df1 = set(df1[id_col])
    df2_filtered = df2[~df2[id_col].isin(ids_df1)]
    
    # Filtrage des colonnes vides ou remplies uniquement de NaN dans df2_filtered
    df2_filtered = df2_filtered.dropna(axis=1, how='all')
    
    # Fusion des deux DataFrames
    df_merged = pd.concat([df1, df2_filtered], ignore_index=True)

    return df_merged

import pandas as pd

def mettre_a_jour_classification(budget_mensuel_donnees, budget_mensuel_categories):
    # Filtrer les lignes où la classification n'est pas NaN ou vide
    filtered_data = budget_mensuel_categories.loc[
        budget_mensuel_categories['Classification'].notna() & 
        (budget_mensuel_categories['Classification'] != '')
    ].copy()  # Ajout de .copy() pour éviter SettingWithCopyWarning

    for i in range(len(filtered_data)):
        # Extraire la ligne à copier en tant que DataFrame 
        ligne_a_copier = filtered_data.iloc[[i]].copy()  # Ajout de .copy()

        # Trouver l'index correspondant dans budget_mensuel_donnees
        index_ligne_donnees = budget_mensuel_donnees[
            budget_mensuel_donnees['ID'] == ligne_a_copier.iloc[0]['ID']
        ].index

        # Vérifier si un index correspondant est trouvé et mettre à jour la ligne
        if not index_ligne_donnees.empty:
            index_cible = index_ligne_donnees[0]

            # Assurer la compatibilité des types avant mise à jour
            for col in ligne_a_copier.columns:
                if col in budget_mensuel_donnees.columns:
                    if pd.api.types.is_numeric_dtype(budget_mensuel_donnees[col]) and pd.api.types.is_numeric_dtype(ligne_a_copier[col]):
                        ligne_a_copier[col] = ligne_a_copier[col].astype(budget_mensuel_donnees[col].dtype)
                    elif budget_mensuel_donnees[col].dtype == 'object' and pd.api.types.is_string_dtype(ligne_a_copier[col]):
                        ligne_a_copier[col] = ligne_a_copier[col].astype('object')

            # Mettre à jour la ligne dans budget_mensuel_donnees
            budget_mensuel_donnees.loc[index_cible] = ligne_a_copier.iloc[0]

    # Mettre à jour budget_mensuel_categories avec les lignes où 'Classification' est NaN ou vide
    budget_mensuel_categories = budget_mensuel_donnees.loc[
        budget_mensuel_donnees['Classification'].isna() | 
        (budget_mensuel_donnees['Classification'] == '')
    ].copy()

    return budget_mensuel_donnees, budget_mensuel_categories


def tri_par_semaine(df):
    # S'assurer qu'on travaille sur une copie explicite pour éviter les avertissements
    df = df.copy()

    # Créer la colonne 'Début de Semaine' avec le premier jour de chaque semaine
    df['Début de Semaine'] = df['Date operation'].dt.to_period('W').dt.start_time.dt.date
    
    # Créer la colonne 'Fin de Semaine' avec le dernier jour de chaque semaine
    df['Fin de Semaine'] = (df['Début de Semaine'] + pd.Timedelta(days=6))

    # Créer la colonne 'Semaine' avec la plage de dates "Début de Semaine - Fin de Semaine"
    df['Semaine'] = df['Début de Semaine'].astype(str) + ' - ' + df['Fin de Semaine'].astype(str)
    
    # Trier le DataFrame par la colonne 'Semaine'
    df = df.sort_values(by='Semaine')
    
    # Retourner le DataFrame avec la colonne 'Semaine' ajoutée
    return df


def calcul_et_tri(df):
    # S'assurer qu'on travaille sur une copie explicite pour éviter les avertissements
    df = df.copy()

    # Filtrer les lignes où 'Classification' appartient à la liste spécifiée
    classifications_valide = [
        "Courses", "Snacks", "Restaurants", "Sport", "Vêtements/Coiffure", 
        "Loisirs", "Divers", "Commande Internet", "Transports", "Autre 1", "Autre 2"
    ]
    df_filtre = df[df['Classification'].isin(classifications_valide)].copy()  # Filtrer uniquement les classifications valides
    
    # Créer la colonne 'Début de Semaine' avec le premier jour de chaque semaine
    df_filtre['Début de Semaine'] = df_filtre['Date operation'].dt.to_period('W').dt.start_time.dt.date
    
    # Créer la colonne 'Fin de Semaine' avec le dernier jour de chaque semaine
    df_filtre['Fin de Semaine'] = (df_filtre['Début de Semaine'] + pd.Timedelta(days=6))

    # Créer la colonne 'Semaine' avec la plage de dates "Début de Semaine - Fin de Semaine"
    df_filtre['Semaine'] = df_filtre['Début de Semaine'].astype(str) + ' - ' + df_filtre['Fin de Semaine'].astype(str)

    # Garder les lignes restantes qui ne sont pas dans les classifications valides
    df_non_filtre = df[~df['Classification'].isin(classifications_valide)].copy()
    
    # Ajouter une colonne 'Semaine' vide pour les lignes non filtrées
    df_non_filtre['Semaine'] = None
    
    # Fusionner les DataFrames filtrés et non filtrés
    df = pd.concat([df_filtre, df_non_filtre], ignore_index=True)

    # Vérifier si les colonnes 'Debit' et 'Credit' existent dans le DataFrame
    if 'Debit' not in df.columns or 'Credit' not in df.columns:
        raise KeyError("Les colonnes 'Debit' et 'Credit' ne sont pas présentes dans le DataFrame.")
    
    # Calculer la somme des colonnes "Debit" et "Credit" pour chaque combinaison de 'Semaine' et 'Classification'
    df_somme = df[df['Classification'].isin(classifications_valide)].groupby(['Semaine', 'Classification'])[['Debit', 'Credit']].sum().reset_index()
    
    # Ajouter une colonne 'Total' qui est la somme de 'Debit' et 'Credit'
    df_somme['Total'] = df_somme['Debit'] + df_somme['Credit']
    
    # Garder uniquement les colonnes 'Semaine', 'Classification' et 'Total'
    df_somme = df_somme[['Semaine', 'Classification', 'Total']]
    
    # DataFrames pour les autres classifications (non valides)
    data_charges_exceptionnelles = []
    data_charges_fixes = []
    data_revenus_exceptionnels = []
    data_revenus_fixes = []
    data_virement_interne = []
    
    # Ajouter les nouvelles classifications aux listes correspondantes
    classifications_charges_exceptionnelles = [
        "Charges exceptionnelles"
    ]
    classifications_charges_fixes = [
        "Trade Républic", "Electricité & Gaz", "Spotify & Apple Storage",
    ]
    classifications_revenus_exceptionnels = [
        "Revenu Exceptionnel"
    ]
    classifications_revenus_fixes = [
        "Bourses"
    ]
    classifications_virement_interne = ['Virement interne' ]
    
    # Créer les listes pour les autres DataFrames
    for _, row in df.iterrows():
        classification = row['Classification']
        
        if classification in classifications_charges_exceptionnelles:
            data_charges_exceptionnelles.append(row)
        elif classification in classifications_charges_fixes:
            data_charges_fixes.append(row)
        elif classification in classifications_revenus_exceptionnels:
            data_revenus_exceptionnels.append(row)
        elif classification in classifications_revenus_fixes:
            data_revenus_fixes.append(row)
        elif classification in classifications_virement_interne:
            data_virement_interne.append(row)
    # Convertir les listes en DataFrames
    data_charges_exceptionnelles = pd.DataFrame(data_charges_exceptionnelles)
    data_charges_fixes = pd.DataFrame(data_charges_fixes)
    data_revenus_exceptionnels = pd.DataFrame(data_revenus_exceptionnels)
    data_revenus_fixes = pd.DataFrame(data_revenus_fixes)
    data_virement_interne= pd.DataFrame(data_virement_interne)
    
    # Supprimer les 3 dernières colonnes de chaque DataFrame
    data_charges_exceptionnelles = data_charges_exceptionnelles.iloc[:, :-3]
    data_charges_fixes = data_charges_fixes.iloc[:, :-3]
    data_revenus_exceptionnels = data_revenus_exceptionnels.iloc[:, :-3]
    data_revenus_fixes = data_revenus_fixes.iloc[:, :-3]
    data_virement_interne = data_virement_interne.iloc[:, :-3]
    # Retourner df_somme et les autres DataFrames
    return df_somme, data_charges_exceptionnelles, data_charges_fixes, data_revenus_exceptionnels, data_revenus_fixes, data_virement_interne



def envoie_donnees(df, file_path):
    nom_feuille = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    mois_dict = {i+1: nom_feuille[i] for i in range(12)}
    classifications_depenses = [
        "Courses", "Snacks", "Restaurants", "Sport", "Vêtements/Coiffure", 
        "Loisirs", "Divers", "Commande Internet", "Transports", "Autre 1", "Autre 2"
    ]
    
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Le fichier {file_path} n'a pas été trouvé.")
    
    # Définir les bordures
    border_thick_left_right = Border(
        left=Side(style="medium"),
        right=Side(style="medium")
    )
    border_thick_all_sides = Border(
        top=Side(style="medium"),
        left=Side(style="medium"),
        bottom=Side(style="medium"),
        right=Side(style="medium")
    )
            
    for ws in wb.worksheets:
        # Défusionner les cellules dans la zone spécifiée
        for row in range(12, ws.max_row + 1):
            for col in [9, 10]:  # Colonnes I et J
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="general", vertical="bottom") 
                # Vérifier si la cellule fait partie d'une plage fusionnée
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        ws.unmerge_cells(str(merged_range))
                        break  
                
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                
                if col == 9:  
                    cell.border = border_thick_left_right
                if col == 10:  
                    cell.border = border_thick_left_right

    semaine_data = {}
    
    for index, row in df.iterrows():
        semaine = row["Semaine"]
        classification = row["Classification"]
        total = row["Total"]
        
        if classification not in classifications_depenses:
            continue
        
        if semaine not in semaine_data:
            semaine_data[semaine] = []
        semaine_data[semaine].append((classification, total))

    # Fusionner la cellule de la semaine avec celle à droite
    for ws in wb.worksheets:
        for row in range(12, ws.max_row + 1):
            week_cell = ws.cell(row=row, column=9)
            right_cell = ws.cell(row=row, column=10)

            if week_cell.value:  
                ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)
                week_cell.alignment = Alignment(horizontal="center", vertical="center")  

    mois_totaux = {}
    
    for semaine, data in semaine_data.items():
        start_date, end_date = semaine.split(" - ")
        start_date = pd.to_datetime(start_date)
        end_date = pd.to_datetime(end_date)
        
        jours_par_mois = {i: 0 for i in range(1, 13)}
        for single_date in pd.date_range(start=start_date, end=end_date):
            mois = single_date.month
            jours_par_mois[mois] += 1
        
        mois_max = max(jours_par_mois, key=jours_par_mois.get)
        mois_nom = mois_dict[mois_max]
        
        if mois_nom not in wb.sheetnames:
            raise ValueError(f"La feuille {mois_nom} n'existe pas dans le fichier.")
        
        ws = wb[mois_nom]
        row_to_insert = 12
        while ws.cell(row=row_to_insert, column=9).value is not None:
            row_to_insert += 1
        
        week_cell = ws.cell(row=row_to_insert, column=9, value=f"{semaine}")
        week_cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        week_cell.border = border_thick_left_right  
        
        row_to_insert += 1
        total_depenses_semaine = 0
        
        for classification, total in data:
            ws.cell(row=row_to_insert, column=9, value=classification).font = Font(bold=False)
            ws.cell(row=row_to_insert, column=10, value=total).font = Font(bold=False)
            ws.cell(row=row_to_insert, column=9).border = border_thick_left_right  
            ws.cell(row=row_to_insert, column=10).border = border_thick_left_right  
            total_depenses_semaine += total
            row_to_insert += 1
        
        ws.cell(row=row_to_insert, column=9, value="TOTAL DEPENSES SEMAINE").font = Font(bold=True)
        ws.cell(row=row_to_insert, column=10, value=total_depenses_semaine).font = Font(bold=True)
        ws.cell(row=row_to_insert, column=9).border = border_thick_left_right  
        ws.cell(row=row_to_insert, column=10).border = border_thick_left_right  
        row_to_insert += 1
        
        if mois_nom not in mois_totaux:
            mois_totaux[mois_nom] = 0
        mois_totaux[mois_nom] += total_depenses_semaine
    
    for mois_nom, total_mois in mois_totaux.items():
        ws = wb[mois_nom]
        row_to_insert = 12
        while ws.cell(row=row_to_insert, column=9).value is not None:
            row_to_insert += 1
        
        ws.cell(row=row_to_insert, column=9, value="TOTAL DEPENSES MOIS").font = Font(bold=True)
        ws.cell(row=row_to_insert, column=10, value=total_mois).font = Font(bold=True)
        ws.cell(row=row_to_insert, column=9).border = border_thick_left_right  
        ws.cell(row=row_to_insert, column=10).border = border_thick_left_right  
        
        ws.cell(row=row_to_insert, column=9).border = border_thick_all_sides
        ws.cell(row=row_to_insert, column=10).border = border_thick_all_sides
        
        row_to_insert += 1  
        
        for row in range(row_to_insert, ws.max_row + 1):
            for col in [9, 10]:  
                cell = ws.cell(row=row, column=col)
                cell.border = Border()  
                
    wb.save(file_path)


def envoi_donnees_revenus_exceptionnels(df, file_path):
    # Charger le fichier Excel
    if df.empty:
        return  # Arrêter immédiatement
    
    wb = load_workbook(file_path)
    
    # Liste des noms des feuilles
    nom_feuille = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    
    # Extraire les mois uniques à partir des dates présentes dans le DataFrame
    mois_a_ouvrir = df['Date operation'].apply(lambda x: x.month).unique()
    
    # Définir les bordures
    border_thick_left_right = Border(
        left=Side(style="medium"),
        right=Side(style="medium")
    )
    
    # Supprimer les valeurs des cellules C25 à C29 pour les feuilles correspondant aux mois présents dans le DataFrame
    for mois in mois_a_ouvrir:
        feuille = nom_feuille[mois - 1]  # Convertir le mois en index de feuille
        ws = wb[feuille]
        for row_num in range(25, 30):  # Parcours des lignes 25 à 29
            cell = ws.cell(row=row_num, column=3)  # Colonne C
            cell.value = None  # Effacer la valeur de la cellule
    
    # Traitement des lignes du DataFrame
    for index, row in df.iterrows():
        date_operation = row['Date operation']
        credit = row['Credit']
        mois = date_operation.month
        feuille = nom_feuille[mois - 1]
        ws = wb[feuille]
        
        # Appliquer les bordures sur toutes les cellules de la plage C25-C29
        for row_num in range(25, 30):
            cell = ws.cell(row=row_num, column=3)
            cell.border = border_thick_left_right  # Applique la bordure gauche/droite à chaque cellule
        
        # Appliquer les bordures supplémentaires sur la première et dernière ligne de la plage
        ws.cell(row=25, column=3).border = Border(
            top=Side(style="medium"),
            left=Side(style="medium"),
            right=Side(style="medium")
        )
        
        ws.cell(row=29, column=3).border = Border(
            bottom=Side(style="medium"),
            left=Side(style="medium"),
            right=Side(style="medium")
        )
        
        # Remplir la première cellule vide avec le crédit
        for row_num in range(25, 30):
            cell = ws.cell(row=row_num, column=3)
            if cell.value is None:  # Dès qu'une cellule vide est trouvée
                cell.value = credit  # Ajouter le crédit
                break  # Arrêter après avoir ajouté la première valeur
    
    # Sauvegarder le fichier modifié
    wb.save(file_path)


def envoi_charges_exceptionnelles(df, file_path):
    # Vérifier si la DataFrame est vide
    if df.empty:
        return  # Arrêt immédiat si df est vide

    # Charger le fichier Excel
    wb = load_workbook(file_path)
    
    # Liste des noms des feuilles
    nom_feuille = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    
    # Extraire les mois uniques présents dans la DataFrame
    mois_a_ouvrir = df['Date operation'].apply(lambda x: x.month).unique()
    
    # Définir les bordures pour la colonne G
    border_thick_left_right = Border(
        left=Side(style="medium"),
        right=Side(style="medium")
    )
    
    # Traiter chaque mois concerné
    for mois in mois_a_ouvrir:
        feuille = nom_feuille[mois - 1]
        ws = wb[feuille]

        # Lire les valeurs actuelles des colonnes F et G pour les lignes 25 à 29
        valeurs_existantes = {}
        for row_num in range(24, 30):
            libelle_f = ws.cell(row=row_num, column=6).value  # Colonne F (Libelle simplifie)
            debit_g = ws.cell(row=row_num, column=7).value  # Colonne G (Débit)
            if libelle_f:  # Ajouter uniquement si la cellule F contient quelque chose
                valeurs_existantes[libelle_f] = debit_g

        # Appliquer les bordures sur la colonne G pour les lignes 25-29
        for row_num in range(24, 30):
            cell = ws.cell(row=row_num, column=7)  # Colonne G
            cell.border = border_thick_left_right  # Applique la bordure gauche/droite

        # Appliquer les bordures supplémentaires sur la première et dernière ligne de la plage
        ws.cell(row=24, column=7).border = Border(
            top=Side(style="medium"),
            left=Side(style="medium"),
            right=Side(style="medium")
        )
        ws.cell(row=29, column=7).border = Border(
            bottom=Side(style="medium"),
            left=Side(style="medium"),
            right=Side(style="medium")
        )

        # Traiter chaque ligne du DataFrame
        for index, row in df.iterrows():
            date_operation = row['Date operation']
            libelle = row['Libelle simplifie']  # Utilisation de "Libelle simplifie" pour F
            debit = row['Debit']  # ✅ Correction ici : utilisation de "Débit" au lieu de "Valeur"
            
            # Vérifier si le libellé existe déjà dans la colonne F
            if libelle in valeurs_existantes:
                if valeurs_existantes[libelle] != debit:
                    # Si la valeur en G est différente, la mettre à jour
                    for row_num in range(24, 30):
                        if ws.cell(row=row_num, column=6).value == libelle:
                            ws.cell(row=row_num, column=7).value = debit
                            break  # Mise à jour effectuée, on sort
            else:
                # Chercher une ligne vide dans la plage 25-29 pour ajouter la nouvelle entrée
                for row_num in range(24, 30):
                    if ws.cell(row=row_num, column=6).value is None:
                        ws.cell(row=row_num, column=6).value = libelle  # Ajouter en F
                        ws.cell(row=row_num, column=7).value = debit  # Ajouter en G
                        break  # Sortir après ajout

    # Sauvegarder le fichier Excel
    wb.save(file_path)



def modif_charges_fixe(df, file_path):
    # Vérifier si la DataFrame est vide
    if df.empty:
        return df  # Retourne le DataFrame inchangé
    
    # Dictionnaire de correspondance pour les modifications
    corrections = {
        "APPLE.COM/BILL": "Spotify & Apple Storage",
        "TRADE REPUBLIC IBAN FRANCE": "Trade Républic",
        "ENGIE": "Electricité & Gaz"
        # Ajoute d'autres corrections ici si nécessaire
    }

    # Appliquer les modifications
    df["Libelle simplifie"] = df["Libelle simplifie"].replace(corrections)
    
    # Gestion spéciale pour Spotify (analyse de la chaîne)
    df["Libelle simplifie"] = df["Libelle simplifie"].apply(
        lambda x: "Spotify & Apple Storage" if "Spotify" in x else x
    )
    return df  # Retourne le DataFrame modifié


def envoi_charges_fixe(df, file_path):
    # Vérifier si la DataFrame est vide
    if df.empty:
        return  # Arrêt immédiat si df est vide

    # Charger le fichier Excel
    wb = load_workbook(file_path)
    
    # Liste des noms des feuilles
    nom_feuille = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    
    # Extraire les mois uniques présents dans la DataFrame
    mois_a_ouvrir = df['Date operation'].apply(lambda x: x.month).unique()
    
    # Définir les bordures pour la colonne G (Débit)
    border_thick_left_right = Border(
        left=Side(style="medium"),
        right=Side(style="medium")
    )
    
    # Traiter chaque mois concerné
    for mois in mois_a_ouvrir:
        feuille = nom_feuille[mois - 1]
        ws = wb[feuille]

        # Lire les valeurs actuelles des colonnes F et G pour les lignes 13 à 20
        valeurs_existantes = {}
        for row_num in range(13, 21):  # 21 exclu pour aller jusqu'à 20
            libelle_f = ws.cell(row=row_num, column=6).value  # Colonne F (Libelle simplifie)
            debit_g = ws.cell(row=row_num, column=7).value  # Colonne G (Débit)
            if libelle_f and libelle_f != "Internet":  # Ignorer "Internet"
                valeurs_existantes[libelle_f] = debit_g

        # Appliquer les bordures sur la colonne G pour les lignes 13-20
        for row_num in range(13, 21):
            cell = ws.cell(row=row_num, column=7)  # Colonne G
            cell.border = border_thick_left_right  # Applique la bordure gauche/droite

        # Appliquer les bordures supplémentaires sur la première et dernière ligne de la plage
        ws.cell(row=13, column=7).border = Border(
            top=Side(style="medium"),
            left=Side(style="medium"),
            right=Side(style="medium")
        )
        ws.cell(row=20, column=7).border = Border(
            bottom=Side(style="medium"),
            left=Side(style="medium"),
            right=Side(style="medium")
        )

        # Calculer la somme Debit + Credit par Classification
        somme_par_classification = df.groupby("Classification")[["Debit", "Credit"]].sum().sum(axis=1)
        
        # Traiter chaque ligne du DataFrame
        for index, row in df.iterrows():
            date_operation = row['Date operation']
            libelle = row['Libelle simplifie']  # Utilisation de "Libelle simplifie" pour F
            
            # Remplacement de row['Debit'] par la somme Debit + Credit de sa Classification
            valeur_a_inscrire = somme_par_classification[row["Classification"]]
        
            # Vérifier si le libellé existe déjà dans la colonne F et n'est pas "Internet"
            if libelle in valeurs_existantes:
                # Mettre à jour la valeur en G si elle est différente
                for row_num in range(13, 21):
                    if ws.cell(row=row_num, column=6).value == libelle:
                        ws.cell(row=row_num, column=7).value = valeur_a_inscrire  # Met à jour la valeur en G
                        break  # Une fois mis à jour, on sort de la boucle

    # Sauvegarder le fichier Excel
    wb.save(file_path)




def envoi_revenus_fixes(df, file_path):
    # Vérifier si le DataFrame est vide
    if df.empty:
        return  # Arrêt immédiat si df est vide
    
    # Charger le fichier Excel
    wb = load_workbook(file_path)
    
    # Liste des noms des feuilles (mois)
    nom_feuille = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                   "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

    # Définir les bordures pour la colonne C
    border_thick_left_right = Border(
        left=Side(style="medium"),
        right=Side(style="medium")
    )

    # 🔹 **Parcourir les mois présents dans df**
    mois_a_ouvrir = df['Date operation'].dt.month.unique()

    for mois in mois_a_ouvrir:
        feuille = nom_feuille[mois - 1]  # Sélectionner la feuille correspondante
        
        # Vérifier si la feuille existe dans le fichier Excel
        if feuille in wb.sheetnames:
            ws = wb[feuille]

            # **🔸 Étape 1 : Effacer les anciennes valeurs entre 13 et 20**
            for row_num in range(13, 21):
                ws[f"C{row_num}"].value = None  # Efface la valeur existante
                ws[f"C{row_num}"].border = None  # Supprime la bordure existante

            # **🔸 Étape 2 : Insérer les nouvelles valeurs de "Bourses"**
            for index, row in df[df['Classification'] == "Bourses"].iterrows():
                # Trouver la première ligne vide entre 13 et 20
                for row_num in range(13, 21):
                    if ws[f"C{row_num}"].value is None:  # Vérifie si la cellule est vide
                        ws[f"C{row_num}"].value = row['Credit']  # Insère la valeur
                        ws[f"C{row_num}"].border = border_thick_left_right  # Applique la bordure
                        break  # Sort de la boucle après l'insertion

    # Sauvegarder les modifications dans le fichier Excel
    wb.save(file_path)



def envoi_virement_interne(df, file_path):
    # Vérifier si le DataFrame est vide
    if df.empty:
        return  # Arrêt immédiat si df est vide
    
    # Charger le fichier Excel
    wb = load_workbook(file_path)
    
    # Liste des noms des feuilles (mois)
    nom_feuille = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                   "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

    
    # Définir les bordures pour la colonne C
    border_thick_left_right = Border(
        left=Side(style="medium"),
        right=Side(style="medium")
    )
    
    # Lire les revenus fixes à partir de la feuille Excel
    for index, row in df.iterrows():
        # Extraire le mois à partir de 'Date operation'
        mois = row['Date operation'].month
        feuille = nom_feuille[mois - 1]  # Sélectionner la feuille correspondante
        
        # Vérifier si la feuille existe dans le fichier Excel
        if feuille in wb.sheetnames:
            ws = wb[feuille]
            
            # Vérifier la classification et inscrire la valeur dans la cellule appropriée
            if row['Classification'] == "Vire":
                # Inscrire la valeur dans C15
                ws["C15"].value = row['Credit']
                # Appliquer une bordure à la cellule C15
                ws["C15"].border = border_thick_left_right
            if row['Classification'] == "Bourses":
                # Inscrire la valeur dans C15
                ws["C13"].value = row['Credit']
                # Appliquer une bordure à la cellule C15
                ws["C13"].border = border_thick_left_right
            # Si vous avez d'autres classifications à gérer, vous pouvez les ajouter ici
            # elif row['Classification'] == "Bourses":
            #     ws["C17"].value = row['Credit']
            #     ws["C17"].border = border_thick_left_right
    
    # Sauvegarder les modifications dans le fichier Excel
    wb.save(file_path)


from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def ajouter_liste_deroulante_categories(file_path):
    # Charger le fichier et sélectionner la feuille 'Categories'
    wb = load_workbook(file_path)
    
    if 'Categories' not in wb.sheetnames:
        print("❌ La feuille 'Categories' n'existe pas dans le fichier.")
        return
    
    ws = wb['Categories']

    # Vérifier s'il y a au moins une ligne de données (au-delà de l'en-tête)
    if ws.max_row <= 1:
        print("⚠️ Aucun contenu trouvé (seulement la ligne 1). Aucune modification effectuée.")
        return  # Arrêter la fonction immédiatement

    # Définir 'Categories' comme la feuille active
    wb.active = wb.index(ws)

    # Déterminer la dernière ligne remplie dans la colonne I
    last_row = ws.max_row

    # Définir les catégories avec des séparateurs
    categories = '"Courses,Snacks,Restaurants,Sport,Vêtements/Coiffure,Loisirs,Divers,Commande Internet,Transports,Autre 1,Autre 2,-----,Bourses,Trade Républic,Spotify & Apple Storage,Electricité & Gaz,Virement interne,Crédit,Revenu Exceptionnel,Charges exceptionnelles"'

    # Créer l'objet de validation de données
    dv = DataValidation(type="list", formula1=categories, allow_blank=True)

    # Configurer les messages d'erreur et d'information
    dv.error = 'Votre sélection doit être dans la liste'
    dv.errorTitle = 'Entrée invalide'
    dv.prompt = 'Veuillez choisir une catégorie dans la liste'
    dv.promptTitle = 'Sélection de liste'

    # Appliquer la validation aux cellules de la colonne I jusqu'à la dernière ligne
    ws.add_data_validation(dv)
    dv.add(f'I2:I{last_row}')

    # Enregistrer le fichier
    wb.save(file_path)
    print(f"✅ Modifications enregistrées dans {file_path}")




def envoi_virement_interne(df, file_path):
    # Vérifier si le DataFrame est vide
    if df.empty:
        return  # Arrêt immédiat si df est vide
    
    # Charger le fichier Excel
    wb = load_workbook(file_path)
    
    # Liste des noms des feuilles (mois)
    nom_feuille = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                   "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

    # Définir les styles de bordures
    border_none = Border()  # Supprimer toutes les bordures
    border_thick_sides = Border(left=Side(style="medium"), right=Side(style="medium"))
    border_thick_all = Border(left=Side(style="medium"), right=Side(style="medium"), bottom=Side(style="medium"), top=Side(style="medium"))

    # Trouver les mois présents dans df
    mois_a_ouvrir = df['Date operation'].dt.month.unique()

    # 🔹 **ÉTAPE 1 : Suppression des anciennes valeurs et bordures après la ligne 35**
    for mois in mois_a_ouvrir:
        feuille = nom_feuille[mois - 1]  # Convertir le mois en index de feuille
        if feuille in wb.sheetnames:
            ws = wb[feuille]

            # Suppression des anciennes valeurs après la ligne 35
            for row in range(36, ws.max_row + 1):  # On supprime depuis 36 jusqu'à la fin
                ws[f"B{row}"].value = None
                ws[f"C{row}"].value = None
                ws[f"B{row}"].border = border_none
                ws[f"C{row}"].border = border_none
                ws[f"B{row}"].font = Font(bold=False)
                ws[f"C{row}"].font = Font(bold=False)

    # 🔹 **ÉTAPE 2 : Ajout des nouveaux virements internes après la ligne 35**
    for index, row in df.iterrows():
        mois = row['Date operation'].month
        feuille = nom_feuille[mois - 1]
        
        if feuille in wb.sheetnames:
            ws = wb[feuille]
            
            # Trouver la première ligne vide à partir de la ligne 36
            ligne = 36
            while ws[f"B{ligne}"].value is not None:
                ligne += 1
            
            # Insérer les données en colonne B et C
            ws[f"B{ligne}"] = row["Libelle simplifie"]
            ws[f"C{ligne}"] = row["Credit"]
            
            # Appliquer les bordures sur B et C
            ws[f"B{ligne}"].border = border_thick_sides
            ws[f"C{ligne}"].border = border_thick_sides

    # 🔹 **ÉTAPE 3 : Ajout de la ligne TOTAL VIREMENT INTERNE après les nouvelles données**
    if not df.empty:
        derniere_ligne = ligne
        
        # Ajouter une ligne vide avec bordures après les données ajoutées
        derniere_ligne += 1
        ws[f"B{derniere_ligne}"].border = border_thick_sides
        ws[f"C{derniere_ligne}"].border = border_thick_sides

        # Ligne TOTAL VIREMENT INTERNE
        derniere_ligne += 1
        ws[f"B{derniere_ligne}"] = "TOTAL VIREMENT INTERNE"
        ws[f"B{derniere_ligne}"].font = Font(bold=True)
        ws[f"B{derniere_ligne}"].border = border_thick_all

        # Recalculer le total uniquement pour les lignes après 35
        total_virement = sum(
            ws[f"C{row}"].value for row in range(36, derniere_ligne) if isinstance(ws[f"C{row}"].value, (int, float))
        )

        ws[f"C{derniere_ligne}"] = total_virement
        ws[f"C{derniere_ligne}"].font = Font(bold=True)
        ws[f"C{derniere_ligne}"].border = border_thick_all
    
    # Sauvegarder les modifications
    wb.save(file_path)




import openpyxl
import openpyxl

def reglage_affichage(file_path):
    # Charger le fichier Excel
    wb = openpyxl.load_workbook(file_path)
    
    # Vérifier si la feuille 'Categories' existe
    if 'Categories' not in wb.sheetnames:
        print("❌ La feuille 'Categories' n'existe pas dans le fichier.")
        return
    
    ws = wb['Categories']
    
    # Vérifier s'il y a au moins une ligne de données (au-delà de l'en-tête)
    if ws.max_row <= 1:
        print("⚠️ Aucun contenu trouvé (seulement la ligne 1). Aucune modification effectuée.")
        return  # Arrêter la fonction immédiatement
    
    # Définir la largeur des colonnes spécifiques
    column_widths = {
        'ID': 20,
        'Classification': 20,
        'Date operation': 18,
        'Sous categorie': 24,
        'Libelle simplifie': 65
    }
    
    # Trouver les indices des colonnes
    header_row = ws[1]
    col_indices = {cell.value: cell.column_letter for cell in header_row if cell.value in column_widths}
    
    # Appliquer les largeurs définies
    for col_name, width in column_widths.items():
        if col_name in col_indices:
            ws.column_dimensions[col_indices[col_name]].width = width
    
    # Sauvegarder les modifications
    wb.save(file_path)
    print("✅ Ajustement des colonnes terminé.")





def enregistrement(data_cp, path_data, budget_mensuel_categories, budget_mensuel_donnees, file_path):
    # Vérifier si 'data_cp' est un entier
    if not isinstance(data_cp, int):
        remove(f".\\{path_data}")
    
    try:
        # Utiliser ExcelWriter pour la feuille 'Categories'
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Réécrire uniquement la feuille 'Categories' modifiée
            budget_mensuel_categories.to_excel(writer, sheet_name='Categories', index=False)

        # Attendre un petit moment pour éviter un conflit potentiel
        time.sleep(1)

        # Utiliser ExcelWriter pour la feuille 'Donnees'
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Réécrire uniquement la feuille 'Donnees' modifiée
            budget_mensuel_donnees.to_excel(writer, sheet_name='Donnees', index=False)

        print("✅ Enregistrement terminé avec succès.")
    
    except Exception as e:
        print(f"❌ Erreur lors de l'enregistrement des données : {e}")

#%%
#On récupère la date du jour
date = date.today()
#Ouverture des nouvelles données à fusionner
data, path_data = ouverture_csv(date)
"""
A vérifier si je laisse le fichier excel Budget Mensuel dans le même dossier que l'exécutable et le reste.
"""
#On vérifier que les dossier existent
verifier_dossiers()

file_path="Budget Mensuel.xlsx"
data2 = read_excel(file_path)

destination_path = f'.\\Précédentes versions\\Budget Mensuel - Version du {date}.xlsx'

# Copier et renommer le fichier
shutil.copy(file_path, destination_path)

budget_mensuel_categories = read_excel(file_path, sheet_name="Categories")
budget_mensuel_donnees = read_excel(file_path, sheet_name="Donnees")

colonnes_attendues = ["Date operation", "Libelle simplifie", "Libelle operation", "Categorie", "Sous categorie", "Debit", "Credit", "ID", "Classification"]


budget_mensuel_donnees = verifier_et_organiser_colonnes(budget_mensuel_donnees, colonnes_attendues)
budget_mensuel_categories = verifier_et_organiser_colonnes(budget_mensuel_categories, colonnes_attendues)


data_cp = creation_data_cp(data, budget_mensuel_donnees)

budget_mensuel_donnees =verification_et_fusion(budget_mensuel_donnees, data_cp)

budget_mensuel_donnees, budget_mensuel_categories = mettre_a_jour_classification(budget_mensuel_donnees, budget_mensuel_categories)


data_calcul = budget_mensuel_donnees.loc[budget_mensuel_donnees['Classification'].notna() & (budget_mensuel_donnees['Classification'] != '')]

data_calcul = tri_par_semaine(data_calcul)

data_somme_semaines, data_charges_exceptionnelles, data_charges_fixes, data_revenus_exceptionnels, data_revenus_fixes, data_virement_interne = calcul_et_tri(data_calcul)

envoie_donnees(data_somme_semaines, file_path)

envoi_donnees_revenus_exceptionnels(data_revenus_exceptionnels, file_path)
envoi_charges_exceptionnelles(data_charges_exceptionnelles, file_path)

data_charges_fixes = modif_charges_fixe(data_charges_fixes, file_path)
envoi_charges_fixe(data_charges_fixes, file_path)
envoi_revenus_fixes(data_revenus_fixes, file_path)
envoi_virement_interne(data_virement_interne, file_path)

enregistrement(data_cp, path_data, budget_mensuel_categories, budget_mensuel_donnees, file_path)
ajouter_liste_deroulante_categories(file_path)
reglage_affichage(file_path)
print("Excel mis a jour.")  
print("Appuyez sur une touche pour fermer...")










